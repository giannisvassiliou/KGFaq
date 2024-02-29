import json
import os
from openai import OpenAI
import openpyxl


def read_queries_responses_from_folder(folder_path):
    queries_file_path = os.path.join(folder_path, './Results.xlsx')
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(queries_file_path)
    
    # Define variable to read sheet
    dataframe1 = dataframe.active

    # Get the number of rows and columns in the sheet
    num_rows = dataframe1.max_row
    num_cols = dataframe1.max_column

    # Create a dictionary to store ID and SPARQL query
    query_dict = {}

    # Assuming the ID column is in column A and the QUERY column is in column B
    for row in range(2, num_rows + 1):  # Start from row 2 to skip header
        query_id = dataframe1.cell(row=row, column=1).value
        sparql_query = dataframe1.cell(row=row, column=2).value
        query_dict[query_id] = sparql_query

    # Close the workbook
    dataframe.close()
    responses = {}

    for file_name in os.listdir(folder_path+'./6_response_folder/'):
        print(folder_path+'./response_folder/')
        if file_name.endswith(".json"):
            print("ff "+file_name)
            # Extract unique code from the file name
            unique_code = file_name.replace("query_response_", "").replace(".json", "")
            file_path = os.path.join(folder_path, './6_response_folder/'+file_name)
            print(f"FilePAth {file_path}")
            with open(file_path, 'r',encoding='utf-8') as file:
                data = json.load(file)
            # Check if the necessary keys exist
                if "results" in data and "bindings" in data["results"]:
                    if unique_code == "90":
                        responses[unique_code] = data["results"]["bindings"][:6]
                    else:
                        responses[unique_code] = data["results"]["bindings"][:20]
                else:
                    print(f"Warning: 'results' or 'bindings' not found in {file_path}")
    return query_dict,responses
folder_path = r''
output_file_path = os.path.join(folder_path, 'output_results_5.json')
query_dict,responses = read_queries_responses_from_folder(folder_path)
output_results = {}

client = OpenAI(

    api_key="USE YOUR OWN API KEY"
)

# Convert keys in query_dict to strings
query_dict = {str(key): value for key, value in query_dict.items()}

# Loop through each SPARQL query and its corresponding response
for unique_code, sparql_query in query_dict.items():
    # Get the corresponding response for the current query
    subset_bindings = responses.get(unique_code, [])
    # Skip if there is no response for the current query
    if not subset_bindings:
        continue

    # Define the conversation for translation
    translation_conversation = [
        {"role": "system", "content": "You are a language translation assistant."},
        {"role": "user", "content": f"Provide a plain-language translation of the SPARQL query.For the query just only maintain the question excluding details about ordering and limits and subjects and labels:\n\n{sparql_query}"}
    ]

    # Define the conversation for response generation
    response_conversation = [
        {"role": "system", "content": "You are a response generation assistant."},
        {"role": "user", "content": f"Basen on the provided SPARQL-response, provide a plain-language translation for the response as a sentence without bullets and numbered lists.\n\n{subset_bindings}"}
    ]

    # Call OpenAI API for translation
    translation_response = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=translation_conversation,
        temperature=0.2,
        top_p=0.5,
        frequency_penalty=0.5,
        max_tokens=150
    )

    # Extract the translated query
    translated_query = translation_response.choices[0].message.content

    # Call OpenAI API for response generation using the translated query
    response_response = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=response_conversation + [{"role": "user", "content": translated_query}],
        temperature=0.3,
        top_p=0.8,
        frequency_penalty=0.96,
        max_tokens=300
    )

    # Extract and print the generated response
    generated_response = response_response.choices[0].message.content
    print(f"Subset binding: {subset_bindings}")
    print(f"Unique Code: {unique_code}")
    print(f"Original SPARQL Query: {sparql_query}")
    print(f"Translated Query: {translated_query}")
    print(f"Generated Response: {generated_response}")
    print("\n" + "=" * 50 + "\n")  # Just for separating the outputs


    # Store the results in the output dictionary
    output_results[unique_code] = {
        "sparql_query": sparql_query,
        "translated_query": translated_query,
        "generated_response": generated_response
    }
     # Save the updated results to the JSON file
    with open(output_file_path, "w", encoding="utf-8") as output_file:
        json.dump(output_results, output_file, ensure_ascii=False, indent=2)

    print(f"Results for Unique Code {unique_code} saved to {output_file_path}")
print('End of Program')


# read SPARQL queries from FinalValidSPARQL.txt
#     queries_file_path = os.path.join(folder_path, 'FinalValidSPARQL_2.txt')
#     with open(queries_file_path, 'r', encoding='utf-8') as queries_file:
#         lines = queries_file.readlines()

#     #Dictionary to store unique codes and corresponding SPARQL queries
#     queries_dict = {}

#     # Regex pattern to extract the unique code from a line
#     unique_code_pattern = re.compile(r'\b(\d+\.\d+)\b')

#     # Iterate over lines and populate the dictionary
#     current_unique_code = None
#     current_sparql_query = []

#     for line in lines:
#         match = unique_code_pattern.search(line)
#         if match:
#             # Found a line with a unique code
#             if current_unique_code is not None:
#                 # Save the previous SPARQL query with the unique code
#                 queries_dict[current_unique_code] = ''.join(current_sparql_query).strip()

#             # Update the current unique code
#             current_unique_code = match.group(1)
#             # Start a new list for the current SPARQL query without the unique code
#             current_sparql_query = [re.sub(unique_code_pattern, '', line)]
#         else:
#             # Continue building the current SPARQL query without the unique code
#             current_sparql_query.append(line)

#     # Save the last SPARQL query with the unique code
#     if current_unique_code is not None:
#         queries_dict[current_unique_code] = ''.join(current_sparql_query).strip()

#     # Rest of the code...





